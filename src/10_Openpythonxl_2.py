"""
Uso: Load Workbook
Creador: Andrés Hernández Mata
Version: 1.0.0
Python: 3.9.1
Fecha: 30 Junio 2021
"""

import random
from openpyxl import load_workbook

##Cargamos un libro que ya ha sido creado
print ("Se abre un libro\nSi existe una excepcion, por ejemplo, si el libro no existe, salimos del programa")
try:
    libro = load_workbook("Mi excel.xlsx")
except FileNotFoundError:
    print ("El archivo no se encuentra el programa terminará")
    exit()
input()

##Enlistamos las páginas del libro y si hay más páginas, activamos la 1
##"Sheet" es la página 0
##"Mi primera hoja" es la página 1
print ("Los nombres de las hojas son:", libro.sheetnames)
input()

if len(libro.sheetnames)>1:
    libro.active = 1
    hoja = libro.active
else:
    hoja = libro.active
print ("La página activa es:",hoja.title)
input()

##Accedemos a una celda en específico
##De nuevo, se regresa un objeto a la variable celda
##Para manipular el objeto necesitamos métodos o funciones
##El metodo value regresa el valor que contiene la celda
celda = hoja["A1"]
hoja["A1"] = "Direcciones"
print ("Accedemos a la celda A1 con valor:",celda.value)
input()


##Podemos acceder directamente a las celdas
hoja["B1"] = "Nombre del equipo"
##(fila, columna, contenido)
hoja.cell(1,3,"Puertos abiertos")



print ("Los encabezados de las celdas son:")
print (hoja["A1"].value,end = "\t")
print (hoja["B1"].value, end = "\t")
print (hoja["C1"].value)
input()




##Rango de celdas
##Rango es de tipo tupla
rangoA = hoja["A2":"A6"]
rangoB = hoja["B2":"C6"]

print ("Imprimimos los objetos de las celdas A2 a A6")
for celda in rangoA:
    for objeto in celda:
        print (celda,"contiene:",objeto.value)
input()

print ("Imprimimos los objetos de las celdas B2 a C6")
for fila in rangoB:
    for columna in fila:
        print (columna,"contiene:",columna.value)
input()

##Llenado de celdas
count = 1
for fila in range(2,50):
    hoja.cell(fila,1,"192.168.1."+str(count))
    count += 1

for fila in range(2,7):
    hoja.cell(fila,2,"Equipo"+chr(fila+63))

for fila in range (2,7):
    lista = [str(random.randint(22,50)),str(random.randint(22,50)),str(random.randint(22,50))]
    hoja.cell(fila,3,str(lista))


##Impresión en pantalla de celdas
print ("Impresión de todas las celdas")
for fila in hoja.values:
    for valor in fila:
        print (valor,end= "\t")
    print ("\n")



## Guardar el libro
libro.save("Mi excel.xlsx")



