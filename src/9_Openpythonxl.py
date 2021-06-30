"""
Uso: Workbook
Creador: Andrés Hernández Mata
Version: 1.0.0
Python: 3.9.1
Fecha: 30 Junio 2021
"""

from openpyxl import Workbook   #Importamos la clase Workbook (POO)


libro = Workbook()  #Workbook(): crea un libro y lo almacenamos en "libro", que será nuestro libro de excel
pagina = libro.active   #Cada libro se crea siempre con una página, esta pagina será la activa para editar

pagina1 = libro.create_sheet("Pagina1") #Podemos crear páginas nuevas
pagina1.title = "Mi primera pagina"     #Editar los nombres de las páginas (sheet)

print (libro.sheetnames)    #Retorna una lista con los nombres de las páginas


libro.save ("Mi excel.xlsx")    #Guardamos nuestro libro de excel con el nombre


